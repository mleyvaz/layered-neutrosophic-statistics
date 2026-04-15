"""
Parse the raw Google-Forms xlsx with expert T,I,F annotations into an
anonymized long-format CSV (no timestamps, no PII, expert_id = 1..N).

The raw xlsx is NOT committed to the repo. Place it locally and run:
    python parse_expert_xlsx.py  path/to/raw.xlsx

Writes:  exp_expert_long.csv  (the artifact committed to the repo)
"""
import sys, os, csv

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")


def zone_of(T, I, F):
    if T > 0.50 and I < 0.35 and F < 0.30:
        return 'Consensus'
    if I >= 0.35:
        return 'Ambiguity'
    if T > 0.30 and F > 0.30:
        return 'Contradiction'
    return 'Ignorance'


def main():
    if len(sys.argv) < 2:
        sys.exit(f"Usage: python {sys.argv[0]} path/to/raw_form_responses.xlsx")
    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        sys.exit(f"File not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    n_cols = ws.max_column
    n_rows = ws.max_row
    n_hyp = (n_cols - 1) // 3

    # Extract hypothesis labels from the T-column header
    hyp_labels = []
    for h_idx in range(n_hyp):
        header = ws.cell(1, 2 + 3 * h_idx).value or ''
        label = header.replace('T (Verdad):', '').strip().rstrip('?').strip()
        hyp_labels.append(label[:140])

    rows = []
    for r in range(2, n_rows + 1):
        for h_idx in range(n_hyp):
            c_T = 2 + 3 * h_idx
            try:
                T10 = float(ws.cell(r, c_T).value)
                I10 = float(ws.cell(r, c_T + 1).value)
                F10 = float(ws.cell(r, c_T + 2).value)
            except (TypeError, ValueError):
                continue
            rows.append({
                'expert_id': r - 1,
                'hypothesis_id': h_idx + 1,
                'hypothesis': hyp_labels[h_idx],
                'T_0_10': T10, 'I_0_10': I10, 'F_0_10': F10,
                'T': T10 / 10.0, 'I': I10 / 10.0, 'F': F10 / 10.0,
                'zone': zone_of(T10 / 10.0, I10 / 10.0, F10 / 10.0),
                'paraconsistent': int((T10 / 10.0 + F10 / 10.0) > 1.0),
            })

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exp_expert_long.csv')
    with open(out, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"Wrote {len(rows)} rows -> {out}")


if __name__ == '__main__':
    main()
